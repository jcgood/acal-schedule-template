"""schedule/parse.py — Data loading and schedule grid parsing."""

import csv
import re
from collections import defaultdict

from config import ACAL_TALKS_TAB, BANTO3D_TALKS_TAB, POSTER_TAB, MIN_TITLE_LENGTH

# Default paths for optional input files (override in build.py or via params)
_CHAIR_XLSX_DEFAULT = 'Session-chair-planning-MF519.xlsx'
_ZOOM_XLSX_DEFAULT  = 'ACAL_Bantoid_Zoom_master.xlsx'

# Matches a short poster label like "S15", "F3", "Remote", "7" in col A
_POSTER_NUM_RE = re.compile(r'^(Remote|\d+|[A-Za-z]+\d*)$', re.IGNORECASE)

# Matches a time span like "4:30-5", "8:30–10:30 AM", "11:30-12:00 noon"
_TIME_SPAN_RE = re.compile(
    r'(\d{1,2}(?::\d{2})?)[–-]((\d{1,2})(:\d{2})?)(\s+\S+)?'
)

_MONTH_NUMS = {
    'January': 1, 'February': 2, 'March': 3, 'April': 4,
    'May': 5, 'June': 6, 'July': 7, 'August': 8,
    'September': 9, 'October': 10, 'November': 11, 'December': 12,
}


# ---------------------------------------------------------------------------
# Title normalization
# ---------------------------------------------------------------------------

def _title_key(text):
    """Normalize a title for use as a match-map key: first line, stripped, lowercased."""
    return text.split('\n')[0].strip().lower()


# ---------------------------------------------------------------------------
# Time formatting
# ---------------------------------------------------------------------------

def _infer_ampm(h):
    """Infer AM/PM given conference window 8 AM – 6 PM (nothing before 8)."""
    return 'AM' if 8 <= h <= 11 else 'PM'


def normalize_time_span(text):
    """Add :00, replace hyphen with en-dash, and add AM/PM where absent."""
    def _fix(m):
        start  = m.group(1)
        end    = m.group(2)
        suffix = (m.group(5) or '').strip().upper()
        if ':' not in start: start += ':00'
        if ':' not in end:   end   += ':00'
        if suffix in ('AM', 'PM', 'NOON'):
            return f'{start}–{end} {suffix}'
        start_h = int(start.split(':')[0])
        end_h   = int(end.split(':')[0])
        sa = _infer_ampm(start_h)
        ea = _infer_ampm(end_h)
        if sa == 'PM' and ea == 'AM':
            ea = 'PM'
        if sa == ea:
            return f'{start}–{end} {sa}'
        return f'{start} {sa}–{end} {ea}'
    return _TIME_SPAN_RE.sub(_fix, text)


def _parse_minutes(time_str):
    """Parse a bare time like '4:30', '8:30 AM', '12:00' to minutes since midnight."""
    s = time_str.strip().lower()
    pm = 'pm'   in s and 'noon' not in s
    am = 'am'   in s
    s  = re.sub(r'\s*(am|pm|noon)\s*', '', s).strip()
    m  = re.match(r'^(\d{1,2}):(\d{2})', s)
    if not m:
        return None
    h, mn = int(m.group(1)), int(m.group(2))
    if pm and h != 12: h += 12
    if am and h == 12: h  = 0
    return h * 60 + mn


def _parse_header_date(text):
    """Extract ISO date string from 'Thursday, May 21' → '2026-05-21'.

    Update the year constant below when reusing for a new conference.
    """
    CONF_YEAR = 2026
    m = re.search(
        r'(January|February|March|April|May|June|July|August|'
        r'September|October|November|December)\s+(\d+)',
        text
    )
    if m:
        return f'{CONF_YEAR}-{_MONTH_NUMS[m.group(1)]:02d}-{int(m.group(2)):02d}'
    return None


def _infer_time_range(time_slots):
    """Infer overall time range from individual slot times (used for Bantoid sessions)."""
    times = [s['time'] for s in time_slots if s.get('time')]
    if not times:
        return ''

    def _start(span):
        return span.split('–', 1)[0].strip()

    def _end(span):
        parts = span.split('–', 1)
        return parts[1].strip() if len(parts) > 1 else span.strip()

    start = _start(times[0])
    end   = _end(times[-1])
    if not re.search(r'\b(AM|PM)\b', start, re.IGNORECASE):
        m = re.match(r'(\d{1,2}):', start)
        if m:
            start += f' {_infer_ampm(int(m.group(1)))}'
    return f'{start}–{end}'


def _end_time_24h(time_range):
    """Extract end time from a normalized span like '2:15–4:15 PM' → '16:15'."""
    m = re.search(r'(\d{1,2}):(\d{2})\s+(AM|PM|NOON)\s*$', time_range, re.IGNORECASE)
    if not m:
        return None
    h, mn, ampm = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    if ampm == 'NOON':
        return f'12:{mn:02d}'
    if ampm == 'PM' and h != 12: h += 12
    elif ampm == 'AM' and h == 12: h = 0
    return f'{h:02d}:{mn:02d}'


def validate_slot_times(block_time, time_slots, context):
    """Print warnings if per-talk slot times are out of order."""
    prev_t, prev_s = None, None
    for slot in time_slots:
        ts = slot.get('time', '')
        if not ts:
            continue
        start_str = re.split(r'[–\-]', ts)[0].strip()
        t = _parse_minutes(start_str)
        if t is None:
            print(f'  WARNING {context}: cannot parse slot time {ts!r}')
            continue
        if prev_t is not None and t < prev_t:
            print(f'  WARNING {context}: {ts!r} is out of order after {prev_s!r}')
        prev_t, prev_s = t, ts


# ---------------------------------------------------------------------------
# Grid helpers
# ---------------------------------------------------------------------------

def cell(row, idx):
    return row[idx].strip() if idx < len(row) else ''


def is_day_header(text):
    return bool(re.match(
        r'(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)',
        text.strip(),
    ))


def is_session_label(text):
    return bool(re.match(r'Session\s+\d+', text.strip()))


def _plenary_introducer(title, chair_map):
    """Return the introducer for a plenary title like 'Plenary 1: Firstname Lastname'."""
    if not chair_map or 'plenary' not in title.lower():
        return ''
    colon = title.find(':')
    if colon < 0:
        return ''
    speaker_part = title[colon + 1:].strip()
    for word in reversed(speaker_part.lower().split()):
        intro = chair_map.get(f'plenary:{word}', '')
        if intro:
            return intro
    return ''


# ---------------------------------------------------------------------------
# Drive / Sheets data loading
# ---------------------------------------------------------------------------

def list_drive_folder(drive_svc, folder_id):
    """List all (non-trashed) files in a Drive folder; return {name: view_url}."""
    url_map = {}
    page_token = None
    query = f"'{folder_id}' in parents and trashed=false"
    while True:
        resp = drive_svc.files().list(
            q=query, fields='nextPageToken, files(id, name)',
            pageSize=1000, pageToken=page_token,
        ).execute()
        for f in resp.get('files', []):
            url_map[f['name']] = f"https://drive.google.com/file/d/{f['id']}/view"
        page_token = resp.get('nextPageToken')
        if not page_token:
            break
    return url_map


def load_match_map(matches_csv, abstract_url_map):
    """Return dict (tab, normalized_title) -> abstract URL."""
    result = {}
    with open(matches_csv, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            name = row.get('abstract_name', '').strip()
            if name and row.get('strikethrough', 'False') != 'True':
                url = abstract_url_map.get(name)
                if url:
                    key = (row['tab'], _title_key(row.get('schedule_title', '')))
                    result[key] = url
    return result


def load_folder_map(csv_path):
    """Return dict (tab, sheet_row, sheet_col) -> folder URL."""
    result = {}
    try:
        with open(csv_path, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                if row['tab'] and row['row'] and row['col']:
                    key = (row['tab'], int(row['row']), int(row['col']))
                    result[key] = row['url']
    except FileNotFoundError:
        pass
    return result


def sync_session_folder_rows(csv_path, acal_grid, banto_grid, poster_grid=None):
    """Rescan fetched grids and update any stale row numbers in session_folders.csv.

    Matches CSV entries to grid positions by rank: the Nth entry for a given
    (tab, col) maps to the Nth session header found in that tab/column.
    """
    def _is_sess(text):
        return bool(re.match(r'Session\s+\d+', text.split('\n')[0].strip()))

    discovered = defaultdict(list)

    for i, row in enumerate(acal_grid):
        col_a = (row[0] if row else '').strip()
        cols  = [row[j].strip() if len(row) > j else '' for j in range(1, 6)]
        non_empty = [c for c in cols if c]
        if col_a and len(non_empty) > 1 and all(_is_sess(c) for c in non_empty):
            for j, c in enumerate(cols):
                if c:
                    discovered[(ACAL_TALKS_TAB, j + 2)].append(i + 1)

    for i, row in enumerate(banto_grid):
        col_a = (row[0] if row else '').replace('\n', '').strip()
        col_b = row[1].strip() if len(row) > 1 else ''
        if not col_a and is_session_label(col_b):
            discovered[(BANTO3D_TALKS_TAB, 2)].append(i + 1)

    if poster_grid is not None:
        _poster_hdr_re = re.compile(r'poster session', re.IGNORECASE)
        for i, row in enumerate(poster_grid):
            col_a  = (row[0] if row else '').strip()
            others = [row[j].strip() if len(row) > j else '' for j in range(1, 4)]
            if col_a and not any(others) and _poster_hdr_re.search(col_a):
                discovered[(POSTER_TAB, 1)].append(i + 1)

    try:
        with open(csv_path, newline='', encoding='utf-8') as f:
            reader     = csv.DictReader(f)
            fieldnames = reader.fieldnames
            entries    = list(reader)
    except FileNotFoundError:
        return

    groups = defaultdict(list)
    for entry in entries:
        if entry.get('tab') and entry.get('row') and entry.get('col'):
            groups[(entry['tab'], int(entry['col']))].append(entry)
    for g in groups.values():
        g.sort(key=lambda e: int(e['row']))

    updated = 0
    for (tab, col), group in groups.items():
        actual = discovered.get((tab, col), [])
        if len(actual) != len(group):
            print(f'  WARNING: {tab} col {col}: {len(group)} in CSV '
                  f'but {len(actual)} found in sheet — check session_folders.csv.')
            continue
        for idx, entry in enumerate(group):
            if int(entry['row']) != actual[idx]:
                entry['row'] = str(actual[idx])
                updated += 1

    if updated:
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(entries)
        print(f'  Auto-updated {updated} session row(s) in {csv_path}.')


def load_chair_map(xlsx_path=_CHAIR_XLSX_DEFAULT):
    """Load confirmed session chairs from the planning spreadsheet.

    Returns dict: ACAL key '1A'/'2B'/... and Banto3d key 'B1'/'B2'/...
    → chair name.  Only rows marked 'confirmed' are included.
    Also populates '<key>_tech' entries for tech coordinators.
    """
    import pandas as pd
    chair_map = {}
    try:
        xl = pd.ExcelFile(xlsx_path)
    except FileNotFoundError:
        return chair_map

    if 'ACAL' in xl.sheet_names:
        df = xl.parse('ACAL', header=0)
        for _, row in df.iterrows():
            session_id = str(row.iloc[0] or '').strip()
            chair      = str(row.iloc[1] or '').strip()
            tech       = str(row.iloc[2] or '').strip()
            confirmed  = str(row.iloc[3] or '').strip().lower()
            if not session_id or confirmed != 'confirmed':
                continue
            m = re.search(r'(\d+)\s*([A-Z])', session_id, re.IGNORECASE)
            if m:
                key = m.group(1) + m.group(2).upper()
                if chair and chair.lower() != 'nan':
                    chair_map[key] = chair
                if tech and tech.lower() != 'nan':
                    chair_map[key + '_tech'] = tech

    if 'Banto3d' in xl.sheet_names:
        df = xl.parse('Banto3d', header=0)
        for _, row in df.iterrows():
            num   = str(row.iloc[0] or '').strip()
            chair = str(row.iloc[1] or '').strip()
            tech  = str(row.iloc[2] or '').strip()
            if num.isdigit():
                if chair and chair.lower() != 'nan':
                    chair_map[f'B{num}'] = chair
                if tech and tech.lower() != 'nan':
                    chair_map[f'B{num}_tech'] = tech

    if 'Plenary intro' in xl.sheet_names:
        df = xl.parse('Plenary intro', header=0)
        for _, row in df.iterrows():
            speaker    = str(row.iloc[0] or '').strip()
            introducer = str(row.iloc[1] or '').strip()
            if speaker and introducer and introducer.lower() != 'nan':
                for word in speaker.lower().split():
                    chair_map[f'plenary:{word}'] = introducer

    return chair_map


def load_zoom_map(xlsx_path=_ZOOM_XLSX_DEFAULT):
    """Load Zoom links from the master spreadsheet.

    Columns: Session | Rooms | Columns | Zoom Link | Passcode | Meeting ID
    Returns dict: room name -> {"url": ..., "passcode": ..., "meeting_id": ...}.
    """
    import pandas as pd
    zoom_map = {}
    try:
        xl = pd.ExcelFile(xlsx_path)
    except FileNotFoundError:
        return zoom_map
    df = xl.parse('Sheet1', header=0)
    for _, row in df.iterrows():
        rooms_raw  = str(row.iloc[1] if not pd.isna(row.iloc[1]) else '').strip()
        url        = str(row.iloc[3] if not pd.isna(row.iloc[3]) else '').strip()
        passcode   = str(row.iloc[4] if not pd.isna(row.iloc[4]) else '').strip()
        meeting_id = str(row.iloc[5] if not pd.isna(row.iloc[5]) else '').strip()
        if not rooms_raw or not url:
            continue
        rooms = [r.strip() for r in rooms_raw.split(',') if r.strip()]
        # CTRC 5019A alias: schedule uses both spellings for the same Bantoid room
        if 'CTRC 5019' in rooms and 'CTRC 5019A' not in rooms:
            rooms.append('CTRC 5019A')
        for room in rooms:
            zoom_map[room] = {'url': url, 'passcode': passcode, 'meeting_id': meeting_id}
    return zoom_map


def fetch_grid(gc, spreadsheet_id, tab_name):
    ws = gc.open_by_key(spreadsheet_id).worksheet(tab_name)
    return ws.get_all_values()


def load_unregistered_set(xlsx_path='registration_check.xlsx'):
    """Return set of (tab, title_key) for talks/posters where registered == 'NO'."""
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(xlsx_path)
        ws = wb['Not Confirmed'] if 'Not Confirmed' in wb.sheetnames else wb.active
        headers   = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        tab_idx   = headers.index('tab')
        title_idx = headers.index('schedule_title')
        result = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            result.add((row[tab_idx], _title_key(row[title_idx] or '')))
        return result
    except (FileNotFoundError, ValueError, TypeError):
        return set()


def load_struck_set(matches_csv):
    """Return set of (tab, normalized_title) for struck-through talk cells."""
    struck = set()
    with open(matches_csv, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            if row.get('strikethrough', 'False') == 'True':
                struck.add((row['tab'], _title_key(row.get('schedule_title', ''))))
    return struck


def load_remote_cells(sheets_svc, spreadsheet_id, tab_name):
    """Return set of (sheet_row, sheet_col) 1-indexed for cells with Light Orange 3 bg."""
    ORANGE = (0xfc / 255, 0xe5 / 255, 0xcd / 255)
    TOL = 0.02
    range_name = f"'{tab_name}'" if ' ' in tab_name else tab_name
    result_data = sheets_svc.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        ranges=[range_name],
        fields='sheets/data/rowData/values/effectiveFormat/backgroundColor',
    ).execute()
    remote = set()
    for r_idx, row_data in enumerate(
        result_data['sheets'][0]['data'][0].get('rowData', [])
    ):
        for c_idx, cell_data in enumerate(row_data.get('values', [])):
            bg = cell_data.get('effectiveFormat', {}).get('backgroundColor', {})
            r = bg.get('red',   1.0)
            g = bg.get('green', 1.0)
            b = bg.get('blue',  1.0)
            if (abs(r - ORANGE[0]) < TOL and abs(g - ORANGE[1]) < TOL
                    and abs(b - ORANGE[2]) < TOL):
                remote.add((r_idx + 1, c_idx + 1))
    return remote


# ---------------------------------------------------------------------------
# ACAL-format Talks parser
# ---------------------------------------------------------------------------

def parse_acal(grid, tab_name, match_map, folder_map,
               struck_set=None, remote_cells=None, chair_map=None):
    """Parse a multi-track talks grid (ACAL format) into a list of event dicts.

    Session block dicts contain:
      time         – overall time range label
      sessions     – list of {name, room, topic, folder_url, sheet_col}
      time_slots   – list of {time, talks: {sheet_col: talk_dict}}
      session_cols – ordered list of sheet_col values
    """
    def _is_session_name(text):
        return bool(re.match(r'Session\s+\d+', text.split('\n')[0].strip()))

    def _is_new_block(na, non_empty_n):
        if not na:
            return False
        if len(non_empty_n) <= 1:
            return True
        return all(_is_session_name(c) for c in non_empty_n)

    events = []
    i = 0
    pending_session_label = ''
    current_date = None

    while i < len(grid):
        row     = grid[i]
        col_a   = cell(row, 0).replace('\n', '')
        col_b   = cell(row, 1)
        cols_bf = [cell(row, j) for j in range(1, 6)]

        if not col_a and not any(cols_bf):
            i += 1
            continue

        if not col_a and is_day_header(col_b) and not any(cols_bf[1:]):
            current_date = _parse_header_date(col_b)
            events.append({'type': 'day', 'text': col_b})
            i += 1
            continue

        if not col_a and is_session_label(col_b) and not any(cols_bf[1:]):
            pending_session_label = col_b
            i += 1
            continue

        if col_a.startswith('Time '):
            i += 1
            continue
        if not col_a and col_b and not any(cols_bf[1:]) and '\n' not in col_b:
            i += 1
            continue

        non_empty = [c for c in cols_bf if c]
        if col_a and len(non_empty) > 1 and all(_is_session_name(c) for c in non_empty):
            session_header_row = i + 1
            session_items = []
            for j in range(5):
                if cols_bf[j]:
                    parts = cols_bf[j].split('\n', 1)
                    topic = re.sub(r'^Banto3d:', 'Bantoid 3:', parts[1].strip()) if len(parts) > 1 else ''
                    session_items.append((j + 2, parts[0].strip(), topic))
            session_col_set = {sc for sc, _, _ in session_items}
            time_range = normalize_time_span(col_a)
            i += 1

            rooms = []
            chairs = []
            techs = []
            subheader_idx = 0
            time_slots = []

            while i < len(grid):
                nrow       = grid[i]
                na         = cell(nrow, 0).replace('\n', '')
                nb_f       = [cell(nrow, j) for j in range(1, 6)]
                non_empty_n = [c for c in nb_f if c]
                has_nl     = any('\n' in c for c in non_empty_n)

                if _is_new_block(na, non_empty_n):
                    break

                if not na and non_empty_n and not has_nl:
                    if nb_f[0]:
                        if subheader_idx == 0:
                            rooms = nb_f[:]
                        elif subheader_idx == 1:
                            chairs = nb_f[:]
                        elif subheader_idx == 2:
                            techs = nb_f[:]
                        subheader_idx += 1
                    i += 1
                    continue

                if has_nl:
                    slot      = {'time': normalize_time_span(na) if na else na, 'talks': {}}
                    sheet_row = i + 1
                    for j in range(5):
                        c_val = cell(nrow, j + 1)
                        if '\n' not in c_val:
                            continue
                        sheet_col = j + 2
                        if sheet_col not in session_col_set:
                            continue
                        first_nl = c_val.index('\n')
                        author   = c_val[:first_nl].strip()
                        title    = c_val[first_nl + 1:].strip()
                        if len(title) >= MIN_TITLE_LENGTH:
                            slot['talks'][sheet_col] = {
                                'author':       author,
                                'title':        title,
                                'abstract_url': match_map.get((tab_name, _title_key(title))),
                                'strikethrough': (tab_name, _title_key(title)) in (struck_set or set()),
                                'remote':       (sheet_row, sheet_col) in (remote_cells or set()),
                                'unreg_key':    (tab_name, _title_key(title)),
                            }
                    if slot['talks']:
                        time_slots.append(slot)
                    i += 1
                    continue

                i += 1

            validate_slot_times(time_range, time_slots, f'ACAL {time_range}')

            sessions = []
            for sheet_col, name, embedded_topic in session_items:
                room_idx       = sheet_col - 2
                sheet_chair    = chairs[room_idx] if room_idx < len(chairs) else ''
                sheet_tech_raw = techs[room_idx]  if room_idx < len(techs)  else ''
                sheet_tech_val = re.sub(
                    r'^(tech\s*(coordinator)?|chair)\s*:\s*', '',
                    sheet_tech_raw, flags=re.IGNORECASE
                ).strip()
                sheet_tech  = sheet_tech_raw if sheet_tech_val and sheet_tech_val.lower() != 'tbd' else ''
                cm          = re.search(r'(\d+)\s*([A-Z])', name, re.IGNORECASE)
                excel_key   = cm.group(1) + cm.group(2).upper() if cm else ''
                excel_chair = (chair_map or {}).get(excel_key, '')        if excel_key else ''
                excel_tech  = (chair_map or {}).get(excel_key + '_tech', '') if excel_key else ''
                sessions.append({
                    'name':       name,
                    'topic':      embedded_topic,
                    'room':       rooms[room_idx] if room_idx < len(rooms) else '',
                    'chair':      excel_chair or sheet_chair,
                    'tech':       sheet_tech or excel_tech,
                    'folder_url': folder_map.get((tab_name, session_header_row, sheet_col)),
                    'sheet_col':  sheet_col,
                })

            events.append({
                'type':          'session_block',
                'session_label': pending_session_label,
                'time':          time_range,
                'sessions':      sessions,
                'time_slots':    time_slots,
                'session_cols':  [sc for sc, _, _ in session_items],
                'date':          current_date,
            })
            pending_session_label = ''
            continue

        if col_a and col_b:
            parts    = [p.strip() for p in col_b.split('\n') if p.strip()]
            subtitle = parts[1] if len(parts) > 1 else ''
            title    = parts[0] if parts else col_b
            events.append({
                'type':         'event',
                'time':         normalize_time_span(col_a),
                'title':        title,
                'details':      parts[1:],
                'abstract_url': match_map.get((tab_name, _title_key(subtitle))) if subtitle else None,
                'introducer':   _plenary_introducer(title, chair_map),
            })
            i += 1
            continue

        i += 1

    return events


# ---------------------------------------------------------------------------
# Bantoid-format Talks parser (single-column sessions)
# ---------------------------------------------------------------------------

def parse_banto(grid, tab_name, match_map, folder_map,
                struck_set=None, remote_cells=None, chair_map=None):
    """Parse a single-track talks grid (Bantoid format) into a list of event dicts."""
    events = []
    i = 0
    in_session = False
    current_session = None
    current_time_slots = []
    current_date = None

    def flush_session():
        nonlocal current_session, current_time_slots, in_session
        if current_session and current_time_slots:
            validate_slot_times('', current_time_slots, f'Banto {current_session["name"]}')
            events.append({
                'type':         'session_block',
                'time':         _infer_time_range(current_time_slots),
                'sessions':     [current_session],
                'time_slots':   current_time_slots,
                'session_cols': [2],
                'date':         current_date,
            })
        current_session     = None
        current_time_slots  = []
        in_session          = False

    while i < len(grid):
        row   = grid[i]
        col_a = cell(row, 0).replace('\n', '')
        col_b = cell(row, 1)

        if not col_a and not col_b:
            i += 1
            continue

        if not col_a and is_day_header(col_b):
            flush_session()
            current_date = _parse_header_date(col_b)
            events.append({'type': 'day', 'text': col_b})
            i += 1
            continue

        if not col_a and is_session_label(col_b):
            flush_session()
            label_parts = col_b.split('\n', 1)
            banto_num   = re.search(r'(\d+)', label_parts[0])
            banto_chair = (chair_map or {}).get(f'B{banto_num.group(1)}', '')        if banto_num else ''
            banto_tech  = (chair_map or {}).get(f'B{banto_num.group(1)}_tech', '')   if banto_num else ''
            current_session = {
                'name':       label_parts[0].strip(),
                'topic':      label_parts[1].strip() if len(label_parts) > 1 else '',
                'room':       '',
                'chair':      banto_chair,
                'tech':       banto_tech,
                'folder_url': folder_map.get((tab_name, i + 1, 2)),
                'sheet_col':  2,
            }
            banto_subheader_idx = 0
            current_time_slots  = []
            in_session          = True
            i += 1
            continue

        if in_session and current_session is not None:
            if not col_a:
                if '\n' in col_b:
                    first_nl  = col_b.index('\n')
                    author    = col_b[:first_nl].strip()
                    title     = col_b[first_nl + 1:].strip()
                    if len(title) >= MIN_TITLE_LENGTH:
                        sheet_row = i + 1
                        current_time_slots.append({
                            'time': '',
                            'talks': {2: {
                                'author':        author,
                                'title':         title,
                                'abstract_url':  match_map.get((tab_name, _title_key(title))),
                                'strikethrough': (tab_name, _title_key(title)) in (struck_set or set()),
                                'remote':        (sheet_row, 2) in (remote_cells or set()),
                                'unreg_key':     (tab_name, _title_key(title)),
                            }},
                        })
                elif col_b:
                    if banto_subheader_idx == 0:
                        current_session['room'] = col_b
                    elif banto_subheader_idx == 1:
                        if not current_session['chair']:
                            current_session['chair'] = col_b
                    elif banto_subheader_idx == 2:
                        tech_val = re.sub(
                            r'^(tech\s*(coordinator)?|chair)\s*:\s*',
                            '', col_b, flags=re.IGNORECASE
                        ).strip()
                        if tech_val and tech_val.lower() != 'tbd':
                            current_session['tech'] = col_b
                    banto_subheader_idx += 1
                i += 1
                continue

            if col_a:
                if '\n' in col_b:
                    first_nl  = col_b.index('\n')
                    author    = col_b[:first_nl].strip()
                    title     = col_b[first_nl + 1:].strip()
                    if len(title) >= MIN_TITLE_LENGTH:
                        sheet_row = i + 1
                        current_time_slots.append({
                            'time': normalize_time_span(col_a),
                            'talks': {2: {
                                'author':        author,
                                'title':         title,
                                'abstract_url':  match_map.get((tab_name, _title_key(title))),
                                'strikethrough': (tab_name, _title_key(title)) in (struck_set or set()),
                                'remote':        (sheet_row, 2) in (remote_cells or set()),
                                'unreg_key':     (tab_name, _title_key(title)),
                            }},
                        })
                        i += 1
                        continue
                    flush_session()
                else:
                    flush_session()

        if col_a and col_b:
            parts    = [p.strip() for p in col_b.split('\n') if p.strip()]
            title    = parts[0] if parts else col_b
            subtitle = parts[1] if len(parts) > 1 else ''
            events.append({
                'type':         'event',
                'time':         normalize_time_span(col_a),
                'title':        title,
                'details':      parts[1:],
                'abstract_url': match_map.get((tab_name, _title_key(subtitle))) if subtitle else None,
                'introducer':   _plenary_introducer(title, chair_map),
            })
            i += 1
            continue

        i += 1

    flush_session()
    return events


# ---------------------------------------------------------------------------
# Poster Session parser
# ---------------------------------------------------------------------------

def parse_posters(grid, tab_name, match_map, struck_set=None, folder_map=None):
    """Parse Poster Session Assignments grid (table format) into sections."""
    sections = []
    current_section = None

    for i, row in enumerate(grid):
        if i == 0:
            continue  # skip column-header row

        col_a = cell(row, 0)
        col_b = cell(row, 1)
        col_c = cell(row, 2)
        col_d = cell(row, 3)

        if not col_a and not col_b and not col_c and not col_d:
            continue

        # Section header: text only in col A.
        if col_a and not col_b and not col_c and not col_d:
            if _POSTER_NUM_RE.match(col_a.strip()):
                continue  # orphan poster number row (content was deleted)
            if '<br/>' in col_a and current_section is not None:
                current_section['notes'] = col_a
            else:
                current_section = {
                    'header':     col_a,
                    'notes':      '',
                    'posters':    [],
                    'folder_url': (folder_map or {}).get((tab_name, i + 1, 1)),
                }
                sections.append(current_section)
            continue

        # Poster row: title in col D (sheet col 4)
        title   = col_d.strip()
        authors = re.sub(r',\s*,+', ',', col_c.strip().replace('\n', ', ')).strip(', ')
        if title and len(title) >= MIN_TITLE_LENGTH:
            if (tab_name, _title_key(title)) in (struck_set or set()):
                continue  # omit cancelled posters entirely
            if current_section is None:
                current_section = {'header': 'Poster Session', 'posters': []}
                sections.append(current_section)
            current_section['posters'].append({
                'number':       col_a,
                'modality':     col_b,
                'authors':      authors,
                'title':        title,
                'abstract_url': match_map.get((tab_name, _title_key(title))),
                'unreg_key':    (tab_name, _title_key(title)),
            })

    # Sort each section: remote first, all others in sheet order
    for s in sections:
        s['posters'].sort(key=lambda p: 0 if 'remote' in p.get('modality', '').lower() else 1)

    return sections
