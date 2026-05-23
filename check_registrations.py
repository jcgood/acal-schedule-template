"""Cross-reference scheduled presenters against conference registrations.

Usage:
    python check_registrations.py
    python conf.py check

Reads:
    - matches.csv              (scheduled talks/posters with abstract IDs)
    - ACAL_XLSX / BANTO3D_XLSX (author names, corresponding emails, modality)
    - Registration CSV matching REGISTRATIONS_GLOB (most recent date wins)
    - inperson_payments.csv    (optional: cash/bank-transfer payers not in CSV)
    - manual_confirmations.csv (optional: presenters confirmed out-of-band)
    - fuzzy_decisions.csv      (optional: persistent yes/no decisions on fuzzy matches)

Writes:
    - registration_check.xlsx  (four sheets: Confirmed, Fuzzy Match, Not Confirmed,
                                Modality Conflicts)

Matching strategy:
    1. Email match: corresponding email from xlsx vs. any registrant email column
    2. Name fuzzy match (fallback): corresponding author name vs. registrant name

Column names for the registration CSV and abstract xlsx are read from config.py.
Update REGISTRATION_EMAIL_COLS, REGISTRATION_INPERSON_TICKET_COLS,
REGISTRATION_ONLINE_TICKET_COLS, and XLSX_* constants to match your platform.
"""

import csv
import glob
import os
import re

import pandas as pd
from rapidfuzz import fuzz, process

from config import (
    ACAL_XLSX, BANTO3D_XLSX, MATCHES_CSV,
    REGISTRATIONS_GLOB,
    REGISTRATION_EMAIL_COLS,
    REGISTRATION_INPERSON_TICKET_COLS,
    REGISTRATION_ONLINE_TICKET_COLS,
    XLSX_ABSTRACT_NAME_COL,
    XLSX_CORRESPONDING_AUTHOR,
    XLSX_CORRESPONDING_EMAIL,
    XLSX_MODALITY_COL,
    XLSX_AUTHOR_COLS,
)

INPERSON_PAYMENTS_FILE   = 'inperson_payments.csv'
OUTPUT_XLSX              = 'registration_check.xlsx'
NAME_MATCH_THRESHOLD     = 85
FUZZY_DECISIONS_CSV      = 'fuzzy_decisions.csv'
MANUAL_CONFIRMATIONS_CSV = 'manual_confirmations.csv'

OUTPUT_FIELDS = [
    'abstract_name', 'conference', 'tab',
    'schedule_title', 'schedule_author',
    'corresponding_author', 'corresponding_email',
    'xlsx_modality',
    'registered', 'match_method', 'matched_author',
    'registrant_name', 'registrant_email',
    'registration_type',
    'modality_conflict',
    'notes',
    'decision',
]


def latest_registration_csv():
    """Return the registration CSV with the most recent MM.DD.YYYY date in its name."""
    files = glob.glob(REGISTRATIONS_GLOB)
    if not files:
        raise FileNotFoundError(f'No file matching {REGISTRATIONS_GLOB!r} found.')
    def _date_key(path):
        m = re.search(r'(\d{2})\.(\d{2})\.(\d{4})', path)
        return (int(m.group(3)), int(m.group(1)), int(m.group(2))) if m else (0, 0, 0)
    return max(files, key=_date_key)


def registration_type(row):
    """Return 'in-person', 'online', or '' based on which tickets were purchased."""
    for col in REGISTRATION_INPERSON_TICKET_COLS:
        if str(row.get(col, '') or '').strip() not in ('', '0'):
            return 'in-person'
    for col in REGISTRATION_ONLINE_TICKET_COLS:
        if str(row.get(col, '') or '').strip() not in ('', '0'):
            return 'online'
    return ''


def load_registrations():
    """Return (email_to_reg, all_names, pending_payment_emails).

    email_to_reg:          {email: {name, emails, reg_type}}
    all_names:             [(name, email, reg_type)]
    pending_payment_emails: set of emails with unpaid in-person entries
    """
    reg_file = latest_registration_csv()
    print(f'  Registration file: {reg_file}')

    email_to_reg = {}
    all_names    = []

    with open(reg_file, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            emails = set()
            for col in REGISTRATION_EMAIL_COLS:
                val = (row.get(col, '') or '').strip().lower()
                if val:
                    emails.add(val)

            name = (row.get('Attendee Name', '') or '').strip()
            if not name:
                name = (row.get('Registrant Name', '') or '').strip()
            if not name:
                name = (row.get('Full name', '') or '').strip()
            if not name:
                first = (row.get('First Name', '') or '').strip()
                last  = (row.get('Last Name',  '') or '').strip()
                name  = f'{first} {last}'.strip()
            if not name:
                name = (row.get('Primary Registrant', '') or '').strip()

            reg = {'name': name, 'emails': emails, 'reg_type': registration_type(row)}

            for email in emails:
                existing = email_to_reg.get(email)
                if existing and existing['reg_type'] == 'in-person' and reg['reg_type'] == 'online':
                    pass  # keep in-person over online for duplicate registrations
                else:
                    email_to_reg[email] = reg
            if name:
                for email in emails:
                    all_names.append((name, email, reg['reg_type']))
                if not emails:
                    all_names.append((name, '', reg['reg_type']))

    # Merge cash/bank-transfer payers not in the registration CSV
    pending_payment_emails = set()
    if os.path.exists(INPERSON_PAYMENTS_FILE):
        with open(INPERSON_PAYMENTS_FILE, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                email = (row.get('email', '') or '').strip().lower()
                name  = (row.get('name',  '') or '').strip()
                paid  = (row.get('paid',  '') or '').strip().lower() == 'true'
                if not email:
                    continue
                if not paid:
                    pending_payment_emails.add(email)
                email_to_reg[email] = {'name': name, 'emails': {email}, 'reg_type': 'in-person'}
                all_names.append((name, email, 'in-person'))

    return email_to_reg, all_names, pending_payment_emails


def load_xlsx_meta():
    """Return {abstract_name: {corresponding_author, corresponding_email, all_authors, modality}}."""
    meta = {}
    for xlsx_path in [ACAL_XLSX, BANTO3D_XLSX]:
        if not os.path.exists(xlsx_path):
            print(f'  WARNING: {xlsx_path} not found — skipping.')
            continue
        df = pd.read_excel(xlsx_path)
        for _, row in df.iterrows():
            aname = str(row.get(XLSX_ABSTRACT_NAME_COL, '') or '').strip()
            if not aname:
                continue
            all_authors = [
                str(row.get(c, '') or '').strip()
                for c in XLSX_AUTHOR_COLS
                if str(row.get(c, '') or '').strip() not in ('', 'NA', 'nan')
            ]
            corr = str(row.get(XLSX_CORRESPONDING_AUTHOR, '') or '').strip()
            if corr and corr not in all_authors:
                all_authors.insert(0, corr)
            modality = ''
            if XLSX_MODALITY_COL:
                modality = str(row.get(XLSX_MODALITY_COL, '') or '').strip()
            meta[aname] = {
                'corresponding_author': corr,
                'corresponding_email':  str(row.get(XLSX_CORRESPONDING_EMAIL, '') or '').strip().lower(),
                'all_authors':          all_authors,
                'modality':             modality,
            }
    return meta


def load_matches():
    rows = []
    with open(MATCHES_CSV, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            if row['abstract_name'] and row.get('strikethrough', 'False') != 'True':
                rows.append(row)
    return rows


def check_registration(corr_email, corr_author, all_authors, email_to_reg, all_names):
    """Check if any author is registered.

    Returns (status, method, matched_author, reg_name, reg_email, reg_type).
    status: 'YES' (email match), 'FUZZY' (name match — needs review), or 'NO'.
    """
    candidate_names = [n for n, _, _ in all_names]

    if corr_email and corr_email in email_to_reg:
        reg = email_to_reg[corr_email]
        return 'YES', 'email', corr_author or (all_authors[0] if all_authors else ''), reg['name'], corr_email, reg['reg_type']

    for author in all_authors:
        if not author:
            continue
        result = process.extractOne(
            author, candidate_names,
            scorer=fuzz.token_sort_ratio,
            processor=str.lower,
            score_cutoff=NAME_MATCH_THRESHOLD,
        )
        if result:
            matched_name, score, idx = result
            matched_email = all_names[idx][1]
            matched_type  = all_names[idx][2]
            return 'FUZZY', f'name ({score:.0f})', author, matched_name, matched_email, matched_type

    return 'NO', '', '', '', '', ''


def modality_conflict(xlsx_modality, reg_type):
    """Flag mismatches between intended modality (xlsx) and registered ticket type."""
    if not xlsx_modality or not reg_type:
        return ''
    xlsx_online   = 'remote' in xlsx_modality.lower() or 'online' in xlsx_modality.lower()
    xlsx_inperson = 'in-person' in xlsx_modality.lower()
    if xlsx_inperson and reg_type == 'online':
        return 'REGISTERED ONLINE (was in-person)'
    if xlsx_online and reg_type == 'in-person':
        return 'REGISTERED IN-PERSON (was remote)'
    return ''


def load_manual_confirmations():
    """Return {abstract_name: modality} for presenters confirmed out-of-band."""
    result = {}
    if os.path.exists(MANUAL_CONFIRMATIONS_CSV):
        with open(MANUAL_CONFIRMATIONS_CSV, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                aname = row.get('abstract_name', '').strip()
                mod   = row.get('modality', '').strip()
                if aname:
                    result[aname] = mod
    return result


def load_fuzzy_decisions():
    """Return {abstract_name: decision} from CSV and any decisions typed into the xlsx."""
    result = {}
    if os.path.exists(FUZZY_DECISIONS_CSV):
        with open(FUZZY_DECISIONS_CSV, newline='', encoding='utf-8') as f:
            for row in csv.DictReader(f):
                aname = row.get('abstract_name', '').strip()
                dec   = row.get('decision', '').strip().lower()
                if aname and dec:
                    result[aname] = dec
    try:
        import openpyxl as _xl
        wb = _xl.load_workbook(OUTPUT_XLSX)
        for ws in wb.worksheets:
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if 'abstract_name' not in headers or 'decision' not in headers:
                continue
            ai = headers.index('abstract_name')
            di = headers.index('decision')
            for row in ws.iter_rows(min_row=2, values_only=True):
                aname = str(row[ai] or '').strip()
                dec   = str(row[di] or '').strip().lower()
                if aname and dec:
                    result[aname] = dec
    except FileNotFoundError:
        pass
    return result


def save_fuzzy_decisions(decisions):
    with open(FUZZY_DECISIONS_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['abstract_name', 'decision'])
        for aname in sorted(decisions):
            w.writerow([aname, decisions[aname]])


def main():
    print('Loading registrations...')
    email_to_reg, all_names, pending_payment_emails = load_registrations()
    print(f'  {len(email_to_reg)} unique registrant emails loaded.')
    print(f'  {len(all_names)} registrant name entries.')
    if pending_payment_emails:
        print(f'  {len(pending_payment_emails)} in-person payment(s) pending.')

    print('Loading xlsx metadata...')
    xlsx_meta = load_xlsx_meta()

    print(f'Loading matches from {MATCHES_CSV}...')
    matches = load_matches()
    print(f'  {len(matches)} scheduled entries.')

    fuzzy_decisions = load_fuzzy_decisions()
    if fuzzy_decisions:
        print(f'  {len(fuzzy_decisions)} existing fuzzy decision(s) loaded.')

    manual_confirmations = load_manual_confirmations()
    if manual_confirmations:
        print(f'  {len(manual_confirmations)} manual confirmation(s) loaded.')

    results = []

    for m in matches:
        if 'plenary' in m.get('schedule_author', '').lower():
            continue
        aname      = m['abstract_name']
        meta       = xlsx_meta.get(aname, {})
        corr_email = meta.get('corresponding_email', '')
        corr_author = meta.get('corresponding_author', '') or m['schedule_author']
        xlsx_mod   = meta.get('modality', '')
        all_authors = meta.get('all_authors', [corr_author] if corr_author else [])

        status, method, matched_author, reg_name, reg_email, reg_type = check_registration(
            corr_email, corr_author, all_authors, email_to_reg, all_names
        )
        registered = status in ('YES', 'FUZZY')
        conflict   = modality_conflict(xlsx_mod, reg_type) if registered else ''

        notes_parts = []
        if registered and reg_email in pending_payment_emails:
            notes_parts.append('PAYMENT PENDING')
        if registered and matched_author and matched_author != corr_author:
            notes_parts.append(f'CO-AUTHOR PRESENTING: {matched_author}')
        notes = '; '.join(notes_parts)

        if aname in manual_confirmations and status != 'YES':
            status     = 'YES'
            registered = True
            method     = 'manual'
            reg_type   = manual_confirmations[aname] or reg_type
            conflict   = modality_conflict(xlsx_mod, reg_type)
            notes      = 'MANUALLY CONFIRMED'

        decision = fuzzy_decisions.get(aname, '') if status == 'FUZZY' else ''
        if status == 'FUZZY' and decision == 'yes':
            status = 'YES'
        elif status == 'FUZZY' and decision == 'no':
            status     = 'NOT_FOUND'
            registered = False
            conflict   = ''
            notes      = ''
        elif status == 'FUZZY' and decision and decision not in ('yes', 'no'):
            print(f'  WARNING: {aname} has unrecognized decision {decision!r} — expected "yes" or "no"')

        results.append({
            'abstract_name':       aname,
            'conference':          m.get('conference', ''),
            'tab':                 m['tab'],
            'schedule_title':      m['schedule_title'],
            'schedule_author':     m['schedule_author'],
            'corresponding_author': corr_author,
            'corresponding_email':  corr_email,
            'xlsx_modality':       xlsx_mod,
            'registered':          status,
            'match_method':        method,
            'matched_author':      matched_author,
            'registrant_name':     reg_name,
            'registrant_email':    reg_email,
            'registration_type':   reg_type,
            'modality_conflict':   conflict,
            'notes':               notes,
            'decision':            decision,
        })

    all_decisions = {r['abstract_name']: r['decision'] for r in results if r.get('decision')}
    all_decisions.update(fuzzy_decisions)
    if all_decisions:
        save_fuzzy_decisions(all_decisions)

    confirmed     = [r for r in results if r['registered'] == 'YES']
    fuzzy         = [r for r in results if r['registered'] == 'FUZZY']
    not_confirmed = [r for r in results if r['registered'] == 'NO']
    conflicts     = [r for r in results if r['modality_conflict']]

    df_confirmed     = pd.DataFrame(confirmed,     columns=OUTPUT_FIELDS)
    df_fuzzy         = pd.DataFrame(fuzzy,         columns=OUTPUT_FIELDS)
    df_not_confirmed = pd.DataFrame(not_confirmed, columns=OUTPUT_FIELDS)
    df_conflicts     = pd.DataFrame(conflicts,     columns=OUTPUT_FIELDS)

    with pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl') as writer:
        df_confirmed.to_excel(writer,     sheet_name='Confirmed',            index=False)
        df_fuzzy.to_excel(writer,         sheet_name='Fuzzy Match - Review', index=False)
        df_not_confirmed.to_excel(writer, sheet_name='Not Confirmed',        index=False)
        df_conflicts.to_excel(writer,     sheet_name='Modality Conflicts',   index=False)

    print(f'\nResults written to {OUTPUT_XLSX}')
    print(f'  Confirmed:          {len(confirmed)}')
    print(f'  Fuzzy (review):     {len(fuzzy)}')
    print(f'  Not confirmed:      {len(not_confirmed)}')
    print(f'  Modality conflicts: {len(conflicts)}')

    if conflicts:
        print('\nModality conflicts (xlsx vs. registration ticket):')
        for r in conflicts:
            print(f'  {r["modality_conflict"]}')
            print(f'    {r["corresponding_author"]} — {r["schedule_title"][:60]}')

    pending = [r for r in results if r['notes'] == 'PAYMENT PENDING']
    if pending:
        print(f'\n{len(pending)} presenter(s) with in-person payment pending:')
        for r in pending:
            print(f'  {r["corresponding_author"]} <{r["corresponding_email"]}>')

    if fuzzy:
        print(f'\n{len(fuzzy)} fuzzy-matched presenter(s) — review "Fuzzy Match - Review" sheet:')
        for r in fuzzy:
            note = f' | {r["notes"]}' if r['notes'] else ''
            print(f'  {r["corresponding_author"]} → {r["registrant_name"]} <{r["registrant_email"]}> [{r["match_method"]}]{note}')
    if not_confirmed:
        print(f'\n{len(not_confirmed)} unconfirmed presenter(s) — see {OUTPUT_XLSX} for full list.')


if __name__ == '__main__':
    main()
