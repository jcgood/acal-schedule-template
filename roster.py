"""roster.py — generate a unified presenter roster and email lists.

Reads:
    - matches.csv              (scheduled talks/posters with abstract IDs)
    - ACAL_XLSX / BANTO3D_XLSX (author names, corresponding emails, modality)
    - registration_check.xlsx  (optional; if present, adds confirmation status
                                from the 'Not Confirmed' sheet)

Writes:
    - roster.csv               (one row per scheduled presenter)

Usage via conf.py:
    python conf.py roster                     # write roster.csv
    python conf.py roster --unconfirmed       # print unconfirmed presenters
    python conf.py roster --format email      # print name/email list
    python conf.py roster --conference banto  # filter to Banto3d only

Direct usage:
    python roster.py [same flags as above]
"""

import csv
import os
import sys

import openpyxl

from config import (
    ACAL_XLSX, BANTO3D_XLSX, MATCHES_CSV,
    XLSX_ABSTRACT_NAME_COL,
    XLSX_CORRESPONDING_AUTHOR,
    XLSX_CORRESPONDING_EMAIL,
    XLSX_MODALITY_COL,
)

REGISTRATION_CHECK_XLSX = 'registration_check.xlsx'
ROSTER_CSV = 'roster.csv'

ROSTER_FIELDS = [
    'abstract_name', 'conference', 'tab',
    'schedule_author', 'corresponding_author', 'email',
    'modality', 'status',
]


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _load_xlsx_meta():
    """Return {abstract_name: {corresponding_author, email, modality}} from both xlsx files."""
    meta = {}
    for path in [ACAL_XLSX, BANTO3D_XLSX]:
        if not os.path.exists(path):
            print(f'  WARNING: {path} not found — skipping.')
            continue
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            ai = headers.index(XLSX_ABSTRACT_NAME_COL)
            ei = headers.index(XLSX_CORRESPONDING_EMAIL)
            oi = headers.index(XLSX_CORRESPONDING_AUTHOR)
        except ValueError as e:
            print(f'  WARNING: {path} missing expected column: {e}')
            wb.close()
            continue
        mi = headers.index(XLSX_MODALITY_COL) if XLSX_MODALITY_COL and XLSX_MODALITY_COL in headers else -1

        for row in ws.iter_rows(min_row=2, values_only=True):
            aname = str(row[ai] or '').strip()
            if not aname:
                continue
            meta[aname] = {
                'corresponding_author': str(row[oi] or '').strip(),
                'email': str(row[ei] or '').strip().lower(),
                'modality': str(row[mi] or '').strip() if mi >= 0 else '',
            }
        wb.close()
    return meta


def _load_matches():
    """Return list of non-struck match rows from matches.csv."""
    if not os.path.exists(MATCHES_CSV):
        print(f'WARNING: {MATCHES_CSV} not found. Run `python conf.py match` first.')
        return []
    rows = []
    with open(MATCHES_CSV, newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            if row.get('strikethrough', 'False').strip().lower() == 'true':
                continue
            if not row.get('abstract_name', '').strip():
                continue
            rows.append(row)
    return rows


def _load_not_confirmed():
    """Return set of abstract_names in the 'Not Confirmed' sheet (if file exists)."""
    if not os.path.exists(REGISTRATION_CHECK_XLSX):
        return None
    try:
        import pandas as pd
        df = pd.read_excel(REGISTRATION_CHECK_XLSX, sheet_name='Not Confirmed')
        return set(df['abstract_name'].dropna().tolist())
    except Exception as e:
        print(f'  WARNING: could not read {REGISTRATION_CHECK_XLSX}: {e}')
        return None


def _conference_from_abstract(abstract_name):
    aname = abstract_name.lower()
    if 'banto3d' in aname:
        return 'Banto3d'
    return 'ACAL'


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_roster(conference_filter='all'):
    """Return list of roster dicts.

    conference_filter: 'all', 'acal', or 'banto'
    """
    matches = _load_matches()
    xlsx_meta = _load_xlsx_meta()
    not_confirmed_set = _load_not_confirmed()

    seen = set()
    roster = []

    for m in matches:
        aname = m['abstract_name'].strip()
        if aname in seen:
            continue
        seen.add(aname)

        conf = _conference_from_abstract(aname)
        if conference_filter == 'acal' and conf != 'ACAL':
            continue
        if conference_filter == 'banto' and conf != 'Banto3d':
            continue

        meta = xlsx_meta.get(aname, {})
        corr_author = meta.get('corresponding_author', '') or m.get('schedule_author', '')
        email = meta.get('email', '')
        modality = meta.get('modality', '')

        if not_confirmed_set is None:
            status = 'Unknown'
        elif aname in not_confirmed_set:
            status = 'Not Confirmed'
        else:
            status = 'Confirmed'

        roster.append({
            'abstract_name': aname,
            'conference': conf,
            'tab': m.get('tab', ''),
            'schedule_author': m.get('schedule_author', ''),
            'corresponding_author': corr_author,
            'email': email,
            'modality': modality,
            'status': status,
        })

    return roster


def write_roster_csv(roster, path=ROSTER_CSV):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=ROSTER_FIELDS)
        writer.writeheader()
        writer.writerows(roster)
    print(f'{len(roster)} presenter(s) written to {path}')


def print_unconfirmed(roster):
    nc = [r for r in roster if r['status'] == 'Not Confirmed']
    if not nc:
        print('No unconfirmed presenters found.')
        return

    banto = [r for r in nc if r['conference'] == 'Banto3d']
    acal  = [r for r in nc if r['conference'] == 'ACAL']

    if banto:
        print(f'Banto3d unconfirmed ({len(banto)}):')
        for r in banto:
            print(f"  {r['schedule_author']} ({r['abstract_name']})")
    if acal:
        print(f'ACAL unconfirmed ({len(acal)}):')
        for r in acal:
            print(f"  {r['schedule_author']} ({r['abstract_name']})")


def print_email_list(roster):
    """Print name/email pairs grouped by status bucket."""
    buckets = {}
    for r in roster:
        buckets.setdefault(r['status'], []).append(r)

    for status in ('Confirmed', 'Not Confirmed', 'Unknown'):
        if status not in buckets:
            continue
        entries = buckets[status]
        print(f'\n-- {status} ({len(entries)}) --')
        for r in sorted(entries, key=lambda x: (x['corresponding_author'] or x['schedule_author']).lower()):
            name  = r['corresponding_author'] or r['schedule_author']
            email = r['email']
            if email:
                print(f'{name} <{email}>')
            else:
                print(f'{name}  (no email)')


def main(unconfirmed_only=False, output_format='csv', conference='all'):
    roster = build_roster(conference_filter=conference)
    if not roster:
        print('No roster entries found.')
        return

    if unconfirmed_only:
        print_unconfirmed(roster)
    elif output_format == 'email':
        print_email_list(roster)
    else:
        write_roster_csv(roster)


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Generate presenter roster CSV and email lists')
    parser.add_argument('--unconfirmed', action='store_true',
                        help='Print unconfirmed presenters only (no CSV written)')
    parser.add_argument('--format', choices=['csv', 'email'], default='csv',
                        help='Output format (default: csv)')
    parser.add_argument('--conference', choices=['all', 'acal', 'banto'], default='all',
                        help='Filter by conference')
    args = parser.parse_args()
    main(unconfirmed_only=args.unconfirmed, output_format=args.format, conference=args.conference)
